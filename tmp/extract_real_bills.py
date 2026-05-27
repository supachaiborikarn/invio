from __future__ import annotations

import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path("/Users/benzsuphaudphanich/Downloads")
OUT = Path("tmp/extracted_real_bills.json")

THAI_MONTHS = {
    "มกราคม": 1,
    "กุมภาพันธ์": 2,
    "มีนาคม": 3,
    "เมษายน": 4,
    "พฤษภาคม": 5,
    "มิถุนายน": 6,
    "กรกฎาคม": 7,
    "สิงหาคม": 8,
    "กันยายน": 9,
    "ตุลาคม": 10,
    "พฤศจิกายน": 11,
    "ธันวาคม": 12,
}

SHORT_MONTHS = {
    "มค": 1,
    "ม.ค.": 1,
    "กพ": 2,
    "ก.พ.": 2,
    "มีค": 3,
    "มี.ค.": 3,
    "เมษ": 4,
    "เม.ย.": 4,
    "พค": 5,
    "พ.ค.": 5,
    "มิย": 6,
    "มิ.ย.": 6,
    "กค": 7,
    "ก.ค.": 7,
    "สค": 8,
    "ส.ค.": 8,
    "กย": 9,
    "ก.ย.": 9,
    "ตค": 10,
    "ต.ค.": 10,
    "พย": 11,
    "พ.ย.": 11,
    "ธค": 12,
    "ธ.ค.": 12,
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value).replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def invoice_no(value: Any) -> str:
    text = clean(value)
    if ":" in text:
        return clean(text.split(":", 1)[1])
    return text


def iso(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raise ValueError(f"Unsupported date value: {value!r}")


def period_from_thai_text(text: str) -> tuple[str, str] | None:
    for month_name, month in THAI_MONTHS.items():
        match = re.search(rf"{month_name}\s*(\d{{2,4}})", text)
        if match:
            year_text = match.group(1)
            year = int(year_text)
            gregorian = year - 543 if year > 2400 else 2500 + year - 543
            return month_period(gregorian, month)
    for month_name, month in SHORT_MONTHS.items():
        match = re.search(rf"{re.escape(month_name)}\s*(\d{{2,4}})", text)
        if match:
            year = int(match.group(1))
            gregorian = year - 543 if year > 2400 else 2500 + year - 543
            return month_period(gregorian, month)
    return None


def month_period(year: int, month: int) -> tuple[str, str]:
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(year, month + 1, 1) - timedelta(days=1)
    return start.isoformat(), end.isoformat()


def tenant_code(customer_name: str) -> str:
    if "บีเอ็นที" in customer_name:
        return "BNT"
    if "ลาซาด้า" in customer_name:
        return "LAZADA"
    if "แฟลช" in customer_name:
        return "FLASH"
    if "ใต้ฟ้า" in customer_name:
        return "TAIFAH"
    if "ดาวไพศาล" in customer_name:
        return "DAOPAISAAN"
    raise ValueError(f"Unknown customer: {customer_name}")


def item_type(description: str) -> str:
    if "ค่าไฟ" in description or "ไฟฟ้า" in description:
        return "electricity"
    if "ขนส่ง" in description or "เที่ยว" in description:
        return "fuel_transport"
    if "เช่า" in description:
        return "rent"
    return "other"


def invoice_type(items: list[dict[str, Any]]) -> str:
    types = {item["type"] for item in items}
    return next(iter(types)) if len(types) == 1 else "mixed"


def due_from_issue(issue: str) -> str:
    return (date.fromisoformat(issue) + timedelta(days=7)).isoformat()


def parse_utility_workbook(path: Path) -> list[dict[str, Any]]:
    invoices: list[dict[str, Any]] = []
    workbook = load_workbook(path, data_only=False)
    for sheet in workbook.worksheets:
        raw_invoice_no = invoice_no(sheet["H4"].value)
        customer = clean(sheet["B6"].value)
        issue_value = sheet["J7"].value
        if not raw_invoice_no or not customer or not isinstance(issue_value, datetime):
            continue

        previous = number(sheet["E14"].value)
        current = number(sheet["E15"].value)
        items: list[dict[str, Any]] = []
        if previous is not None and current is not None:
            description = clean(sheet["B13"].value)
            qty = int(round(current - previous))
            unit_price = number(sheet["H13"].value) or 0
            items.append(
                {
                    "type": "electricity",
                    "description": description,
                    "quantity": qty,
                    "unitPrice": unit_price,
                    "amount": qty * unit_price,
                    "meter": {
                        "previousReading": int(round(previous)),
                        "currentReading": int(round(current)),
                        "usageUnits": qty,
                        "rate": unit_price,
                    },
                }
            )
        else:
            for row in range(13, 24):
                description = clean(sheet[f"B{row}"].value)
                qty = number(sheet[f"G{row}"].value)
                unit_price = number(sheet[f"H{row}"].value)
                if not description or qty is None or unit_price is None:
                    continue
                items.append(
                    {
                        "type": item_type(description),
                        "description": description,
                        "quantity": int(round(qty)),
                        "unitPrice": unit_price,
                        "amount": qty * unit_price,
                    }
                )
        if not items:
            continue

        first_period = next(
            (period_from_thai_text(item["description"]) for item in items if period_from_thai_text(item["description"])),
            None,
        )
        if first_period is None:
            first_period = month_period(issue_value.year, issue_value.month)

        issue = iso(issue_value)
        invoices.append(
            {
                "source": path.name,
                "sheet": sheet.title,
                "invoiceNo": raw_invoice_no,
                "tenantCode": tenant_code(customer),
                "tenantName": customer,
                "tenantTaxId": clean(sheet["B9"].value).replace("เลขประจำตัวผู้เสียภาษี", "").strip(),
                "tenantAddress": "\n".join(
                    part for part in [clean(sheet["B7"].value), clean(sheet["B8"].value)] if part
                ),
                "type": invoice_type(items),
                "issueDate": issue,
                "dueDate": due_from_issue(issue),
                "periodStart": first_period[0],
                "periodEnd": first_period[1],
                "vatEnabled": False,
                "discount": 0,
                "items": items,
            }
        )
    return invoices


def parse_transport_workbook(path: Path) -> list[dict[str, Any]]:
    invoices: list[dict[str, Any]] = []
    workbook = load_workbook(path, data_only=False)
    for sheet in workbook.worksheets:
        raw_invoice_no = invoice_no(sheet["H3"].value)
        customer = clean(sheet["B5"].value)
        issue_value = sheet["J6"].value
        if not raw_invoice_no or not customer or not isinstance(issue_value, datetime):
            continue
        items: list[dict[str, Any]] = []
        for row in range(12, 28):
            description = clean(sheet[f"B{row}"].value)
            qty = number(sheet[f"G{row}"].value)
            unit_price = number(sheet[f"H{row}"].value)
            if not description or qty is None or unit_price is None:
                continue
            items.append(
                {
                    "type": "fuel_transport",
                    "description": description,
                    "quantity": int(round(qty)),
                    "unitPrice": unit_price,
                    "amount": qty * unit_price,
                }
            )
        if not items:
            continue
        first_period = period_from_thai_text(items[0]["description"]) or month_period(
            issue_value.year, issue_value.month
        )
        issue = iso(issue_value)
        invoices.append(
            {
                "source": path.name,
                "sheet": sheet.title,
                "invoiceNo": raw_invoice_no,
                "tenantCode": tenant_code(customer),
                "tenantName": customer,
                "tenantTaxId": "",
                "tenantAddress": clean(sheet["B6"].value),
                "type": "fuel_transport",
                "issueDate": issue,
                "dueDate": due_from_issue(issue),
                "periodStart": first_period[0],
                "periodEnd": first_period[1],
                "vatEnabled": False,
                "discount": 0,
                "items": items,
            }
        )
    return invoices


def adjusted_year(value: datetime) -> int:
    return value.year + 57 if value.year < 2000 else value.year


def parse_daopaisaan(path: Path) -> list[dict[str, Any]]:
    invoices: list[dict[str, Any]] = []
    workbook = load_workbook(path, data_only=False)

    for sheet_name, start_row, end_row in [("เก่า2", 7, 18), ("สรุป 69", 7, 10)]:
        sheet = workbook[sheet_name]
        for row in range(start_row, end_row + 1):
            month_value = sheet[f"B{row}"].value if sheet_name == "เก่า2" else sheet[f"A{row}"].value
            trips = number(sheet[f"C{row}"].value if sheet_name == "เก่า2" else sheet[f"B{row}"].value)
            liters = number(sheet[f"D{row}"].value if sheet_name == "เก่า2" else sheet[f"C{row}"].value)
            rate = number(sheet[f"E{row}"].value if sheet_name == "เก่า2" else sheet[f"D{row}"].value)
            if not isinstance(month_value, datetime) or trips is None or rate is None:
                continue
            year = adjusted_year(month_value)
            month = month_value.month
            if liters is None:
                liters = trips * 38000
            period_start, period_end = month_period(year, month)
            issue = (date.fromisoformat(period_end) + timedelta(days=1)).isoformat()
            be_year = year + 543
            invoice_no_value = f"DAOPAISAAN-{be_year}-{month:02d}"
            items: list[dict[str, Any]] = []

            if sheet_name == "สรุป 69" and month == 4:
                for detail_row in range(23, 54):
                    day = number(sheet[f"A{detail_row}"].value)
                    detail_liters = number(sheet[f"B{detail_row}"].value)
                    if day is None or detail_liters is None:
                        continue
                    items.append(
                        {
                            "type": "fuel_transport",
                            "description": f"ค่าเที่ยวขนส่งวันที่ {int(day)} เมษายน {be_year}",
                            "quantity": int(round(detail_liters)),
                            "unitPrice": rate,
                            "amount": detail_liters * rate,
                        }
                    )
            elif sheet_name == "เก่า2" and month == 12:
                for detail_row in range(29, 60):
                    day = number(sheet[f"C{detail_row}"].value)
                    detail_liters = number(sheet[f"D{detail_row}"].value)
                    if day is None or detail_liters is None:
                        continue
                    items.append(
                        {
                            "type": "fuel_transport",
                            "description": f"ค่าเที่ยวขนส่งวันที่ {int(day)} ธันวาคม {be_year}",
                            "quantity": int(round(detail_liters)),
                            "unitPrice": rate,
                            "amount": detail_liters * rate,
                        }
                    )

            if not items:
                items.append(
                    {
                        "type": "fuel_transport",
                        "description": f"ค่าเที่ยวขนส่ง {month:02d}/{be_year} จำนวน {int(trips)} เที่ยว",
                        "quantity": int(round(liters)),
                        "unitPrice": rate,
                        "amount": liters * rate,
                    }
                )

            invoices.append(
                {
                    "source": path.name,
                    "sheet": sheet_name,
                    "invoiceNo": invoice_no_value,
                    "tenantCode": "DAOPAISAAN",
                    "tenantName": "ดาวไพศาล",
                    "tenantTaxId": "",
                    "tenantAddress": "",
                    "type": "fuel_transport",
                    "issueDate": issue,
                    "dueDate": due_from_issue(issue),
                    "periodStart": period_start,
                    "periodEnd": period_end,
                    "vatEnabled": False,
                    "discount": 0,
                    "items": items,
                }
            )
    return invoices


def flash_invoice() -> dict[str, Any]:
    return {
        "source": "ใบแจ้งหนี้ flash 12 68.pdf",
        "sheet": "page 1",
        "invoiceNo": "20251203",
        "tenantCode": "FLASH",
        "tenantName": "บริษัท แฟลช เอ็กซ์เพรส จำกัด สำนักงานใหญ่",
        "tenantTaxId": "0105560159254",
        "tenantAddress": "เลขที่ 161 อาคารยูนิลีเวอร์ เฮ้าส์ ชั้นที่ 7 และ 8 ถนนพระรามเก้า\nแขวงห้วยขวาง เขตห้วยขวาง กรุงเทพมหานคร 10310",
        "type": "rent",
        "issueDate": "2025-12-01",
        "dueDate": "2025-12-31",
        "periodStart": "2025-12-01",
        "periodEnd": "2025-12-31",
        "vatEnabled": False,
        "discount": 1552.63,
        "items": [
            {
                "type": "rent",
                "description": "ค่าเช่า (เริ่ม 01/12/2568-31/12/2568)",
                "quantity": 1,
                "unitPrice": 31052.63,
                "amount": 31052.63,
            }
        ],
    }


def main() -> None:
    invoices: list[dict[str, Any]] = []
    invoices.extend(parse_utility_workbook(ROOT / "ค่าไฟ JT วัชรเกียรติ.xlsx"))
    salakabat = ROOT / "ใบแจ้งหนี้ สลกบาตร 2567.xlsx"
    invoices.extend(parse_utility_workbook(salakabat))
    invoices.extend(parse_transport_workbook(salakabat))
    invoices.extend(parse_daopaisaan(ROOT / "ค่าเที่ยวดาวไพศาล.xlsx"))
    invoices.append(flash_invoice())

    OUT.write_text(json.dumps(invoices, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"extracted {len(invoices)} invoices -> {OUT}")
    counts: dict[str, int] = {}
    for invoice in invoices:
        counts[invoice["tenantCode"]] = counts.get(invoice["tenantCode"], 0) + 1
    for code, count in sorted(counts.items()):
        print(f"{code}: {count}")


if __name__ == "__main__":
    main()
