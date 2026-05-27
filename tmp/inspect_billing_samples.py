from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOTS = [
    Path("/Users/benzsuphaudphanich/Downloads/ค่าเที่ยวดาวไพศาล.xlsx"),
    Path("/Users/benzsuphaudphanich/Downloads/ค่าไฟ JT วัชรเกียรติ.xlsx"),
    Path("/Users/benzsuphaudphanich/Downloads/ใบแจ้งหนี้ สลกบาตร 2567.xlsx"),
]

PDFS = [
    Path("/Users/benzsuphaudphanich/Downloads/ใบแจ้งหนี้ flash 12 68.pdf"),
]


def clean(value):
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    while "  " in text:
        text = text.replace("  ", " ")
    return text


def print_sheet(path: Path) -> None:
    print(f"\n=== XLSX: {path.name} ===")
    workbook = load_workbook(path, data_only=False, read_only=False)
    print("sheets:", ", ".join(workbook.sheetnames))

    for sheet in workbook.worksheets:
        print(f"\n--- sheet: {sheet.title} ---")
        print("size:", sheet.max_row, "rows x", sheet.max_column, "cols")
        if sheet.merged_cells.ranges:
            print("merged:", ", ".join(str(item) for item in list(sheet.merged_cells.ranges)[:20]))
        if sheet.print_area:
            print("print_area:", sheet.print_area)
        print("orientation:", sheet.page_setup.orientation, "paper:", sheet.page_setup.paperSize)

        rows_printed = 0
        for row in sheet.iter_rows():
            cells = []
            for cell in row:
                value = clean(cell.value)
                if value:
                    cells.append(f"{cell.coordinate}={value[:90]}")
            if cells:
                print(" | ".join(cells))
                rows_printed += 1
            if rows_printed >= 80:
                print("... truncated ...")
                break


def print_pdf(path: Path) -> None:
    print(f"\n=== PDF: {path.name} ===")
    reader = PdfReader(str(path))
    print("pages:", len(reader.pages))
    for index, page in enumerate(reader.pages[:3], start=1):
        text = clean(page.extract_text() or "")
        media = page.mediabox
        print(f"\n--- page {index} / size {float(media.width):.1f} x {float(media.height):.1f} ---")
        print(text[:4000])


for item in ROOTS:
    print_sheet(item)

for item in PDFS:
    print_pdf(item)
