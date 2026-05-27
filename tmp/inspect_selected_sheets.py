from pathlib import Path

from openpyxl import load_workbook

TARGETS = [
    (
        Path("/Users/benzsuphaudphanich/Downloads/ค่าไฟ JT วัชรเกียรติ.xlsx"),
        ["1.69.1", "5.69.2"],
    ),
    (
        Path("/Users/benzsuphaudphanich/Downloads/ใบแจ้งหนี้ สลกบาตร 2567.xlsx"),
        ["Sheet1", "Sheet4", "Sheet10"],
    ),
]


def clean(value):
    return str(value).replace("\n", " ").strip()[:90]


for path, sheet_names in TARGETS:
    print(f"=== {path.name} ===")
    workbook = load_workbook(path, data_only=False)
    print("sheets:", ", ".join(workbook.sheetnames))
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        print(f"--- {sheet_name} ({sheet.max_row} x {sheet.max_column}) ---")
        for row in range(1, min(sheet.max_row, 45) + 1):
            values = []
            for col in range(1, min(sheet.max_column, 12) + 1):
                cell = sheet.cell(row, col)
                if cell.value not in (None, ""):
                    values.append(f"{cell.coordinate}={clean(cell.value)}")
            if values:
                print(" | ".join(values))
